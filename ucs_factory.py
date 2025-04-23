"""ユースケースシナリオファクトリークラス
"""
import openpyxl
from use_case_scenario import Actor, UseCaseScenario, Condition, FlowType, Flow


def _get_string(cell):
    """セルから文字列を取り出す
    """
    val = cell.value
    if val is None:
        val = ''
    return val


def _create_actors(line):
    """文字列からアクターを生成
    """
    actors = []
    for prim in line.split('、'):
        for prim2 in prim.split('\n'):
            actors.append(Actor(prim2))
    return actors


def _set_header(ws, ucs):
    """ヘッダー情報をセット
    """
    is_pre_cond = False
    is_post_cond = False
    for row in ws.iter_rows():
        val = _get_string(row[2])
        match _get_string(row[0]):
            case 'シナリオID':
                ucs.scenario_id = val
            case '関連要求ID':
                ucs.related_request_id = val
            case '概要、場面':
                ucs.abstract = val
            case 'アクター':
                ucs.actors = _create_actors(val)
            case 'ステークホルダ要求':
                ucs.stakeholder_requirement = val
            case '関連要求（制約）':
                ucs.related_requirement = val
            case '事前条件':
                is_pre_cond = True
                is_post_cond = False
            case '事後条件':
                is_pre_cond = False
                is_post_cond = True
            case 'フロー':
                break

        if is_pre_cond and val:
            # 事前条件処理
            cond = Condition(_get_string(row[1]), val)
            ucs.pre_conditions.append(cond)
        elif is_post_cond and val:
            # 事後条件処理
            cond = Condition(_get_string(row[1]), val)
            ucs.post_conditions.append(cond)


def _set_scenario(ws, ucs):
    """シナリオ（フロー）をセットする
    """
    # フローの先頭まで移動
    iter_rows = ws.iter_rows()  # ジェネレータを取っておく
    for row in iter_rows:
        if _get_string(row[0]) == 'フロー':
            break

    # フロー生成処理
    flow_type = FlowType.NONE
    flow = None
    for row in iter_rows:  # 取っておいたジェネレータを使う
        # ステップ（フローIDの欄）
        r_step = _get_string(row[2])
        # アクション詳細（シナリオ欄）
        r_detail = _get_string(row[3])
        # 分岐欄
        r_branch = _get_string(row[4])
        # 備考欄
        r_note = _get_string(row[5])

        if not r_step:
            continue

        # フロータイプの判定
        match _get_string(row[0]):
            case '基本フロー':
                flow_type = FlowType.MAIN
            case '代替フロー':
                flow_type = FlowType.ALTERNATIVE
            case '例外フロー':
                flow_type = FlowType.EXCEPTION
            case '':
                pass
            case _:
                # 「課題、TBD事項」で終了
                break

        # フロー生成
        if flow_type == FlowType.MAIN:
            # 基本フローの登録
            if flow is None:
                flow = Flow()
                ucs.add_flow(flow)
        elif flow_type == FlowType.ALTERNATIVE:
            # 代替フローの登録
            if '-' not in r_step:
                flow = Flow(FlowType.ALTERNATIVE, r_step)
                ucs.add_flow(flow)
                continue
        elif flow_type == FlowType.EXCEPTION:
            # 例外フローの登録
            if '-' not in r_step:
                flow = Flow(FlowType.EXCEPTION, r_step)
                ucs.add_flow(flow)
                continue

        # アクション登録
        flow.add_action(r_step, r_detail, r_branch, r_note)


def create(excel_file):
    wb = openpyxl.load_workbook(excel_file)
    ws = wb.worksheets[0]
    if ws.title == '変更履歴':
        ws = wb.worksheets[1]
    ucs = UseCaseScenario()
    ucs.excel_path = excel_file

    _set_header(ws, ucs)
    _set_scenario(ws, ucs)

    ucs.traverse_flow()  # アクションのブランチを接続

    return ucs
